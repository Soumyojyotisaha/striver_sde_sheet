from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# Define data
data = {
    "Day 1: Array": [
        "2Sum Problem",
        "Best Time to Buy and Sell Stock",
        "Contains Duplicate",
        "Product of Array Except Self",
        "Kadane's Algorithm",
        "Maximum Product Subarray",
        "Find Minimum in Rotated Sorted Array",
        "Search in Rotated Sorted Array I",
        "3 Sum",
        "Container with Most Water"
    ],
    "Day 2: Binary": [
        "Sum of Two Integers",
        "Number of 1 Bits",
        "Counting Bits",
        "Find Missing Number in an Array",
        "Reverse Bits"
    ],
    "Day 3: Dynamic Programming": [
        "Climbing Stairs",
        "Coin Change",
        "Longest Increasing Subsequence",
        "Longest Common Subsequence",
        "Word Break",
        "Combination Sum",
        "Maximum Sum of Non-Adjacent Elements",
        "House Robber",
        "Decode Ways",
        "Grid Unique Paths",
        "Jump Game"
    ],
    "Day 4: Graph": [
        "Clone Graph",
        "Course Schedule - I",
        "Pacific Atlantic Water Flow",
        "Number of Islands",
        "Longest Consecutive Sequence",
        "Alien Dictionary",
        "Graph Valid Tree",
        "Connected Components"
    ],
    "Day 5: Interval": [
        "Insert Interval",
        "Merge Intervals",
        "Non-Overlapping Intervals",
        "Repeat and Missing Number",
        "Meeting Rooms",
        "Meeting Rooms II"
    ],
    "Day 6: Linked List": [
        "Reverse a LinkedList",
        "Detect a Cycle in Linked List",
        "Merge Two Sorted Linked Lists",
        "Merge K Sorted Arrays",
        "Remove N-th Node from End of Linked List",
        "Reorder List"
    ],
    "Day 7: Matrix": [
        "Set Matrix Zeros",
        "Print the Matrix in Spiral Manner",
        "Rotate Matrix by 90 Degrees",
        "Word Search"
    ],
    "Day 8: String": [
        "Longest Substring Without Repeating Characters",
        "Longest Repeating Character Replacement",
        "Minimum Window Substring",
        "Check for Anagrams",
        "Group Anagrams",
        "Check for Balanced Parentheses",
        "Check Palindrome",
        "Longest Palindromic Substring",
        "Palindromic Substrings",
        "Encode and Decode Strings"
    ],
    "Day 9: Tree": [
        "Height of a Binary Tree",
        "Check if Two Trees Are Identical",
        "Invert/Flip Binary Tree",
        "Maximum Path Sum",
        "Level Order Traversal",
        "Serialize and Deserialize Binary Tree",
        "Subtree of Another Tree",
        "Construct Binary Tree from Postorder and Inorder Traversal",
        "Check if a Tree is a BST",
        "Find K-th Smallest Element in BST",
        "Find LCA of Two Nodes in BST",
        "Implement Trie (Prefix Tree)",
        "Implement Trie - 2"
    ],
    "Day 10: Heap": [
        "K Most Frequent Elements",
        "Find Median from Data Stream"
    ]
}

# Create a workbook and select the active worksheet
wb = Workbook()
ws = wb.active
ws.title = "LeetCode Questions"

# Define header style
header_font = Font(bold=True)
header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Write headers
ws["A1"] = "Day"
ws["B1"] = "Topic"
ws["C1"] = "Question"
for cell in ws["1:1"]:
    cell.font = header_font
    cell.fill = header_fill

# Write data
row = 2
for day, questions in data.items():
    topic = day
    for question in questions:
        ws[f"A{row}"] = day.split(":")[0]
        ws[f"B{row}"] = day.split(":")[1].strip()
        ws[f"C{row}"] = question
        row += 1

# Save the workbook
wb.save("LeetCode_Questions.xlsx")
